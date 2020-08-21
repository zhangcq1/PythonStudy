import openpyxl
def update_data(N):
    wb = openpyxl.load_workbook('tepl.xlsx')
    sh = wb['Sheet1']
    for i in range(3,1002):
        sh.cell(row = i,column = 1,value = N)
        sh.cell(row = i,column = 2,value = N)
        N = N+1
    wb.save('tepl.xlsx')
    wb.close()
update_data(1000)
#1,1000,2000,3000