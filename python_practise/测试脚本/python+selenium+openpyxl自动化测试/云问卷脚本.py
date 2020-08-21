# coding=utf-8🔥
# 1.先设置编码，utf-8可支持中英文，如上，一般放在第一行
# 2.注释：包括记录创建时间，创建人，项目名称。
'''
Created on 2019-12-03
@author: 北京-宏哥   QQ交流群：705269076
Project: python+ selenium自动化测试练习篇4
'''
# 3.导入模块
import time
from selenium import webdriver
import openpyxl


def update_data(N):
    wb = openpyxl.load_workbook('tepl.xlsx')
    sh = wb['Sheet1']
    for i in range(3,1002):
        sh.cell(row = i,column = 1,value = N)
        sh.cell(row = i,column = 2,value = N)
        N = N+1
    wb.save('tepl.xlsx')
    wb.close()




driver = webdriver.Chrome()
driver.maximize_window()
driver.implicitly_wait(6)
driver.get("http://ywj.10085.cn/login.html")
time.sleep(1)
# driver.find_element_by_class_name('title').click()
old_page = driver.current_url
driver.find_element_by_id('phone').send_keys('15690866308')
driver.find_element_by_id('pwd').send_keys('zcq13781240894')
time.sleep(10)
while True:
    time.sleep(2)
    new_page = driver.current_url
    if new_page != old_page:
        break
    print('页面未切换')
driver.find_element_by_xpath('//div[@class="header-user"]').click()
time.sleep(1)
driver.find_element_by_xpath('//ul[@class="menu-list-box"]/li[3]').click()
num = 1
while 1:
    update_data(num)
    time.sleep(3)
    driver.find_element_by_xpath('//label').click()
    time.sleep(5)
    driver.find_element_by_id('upload-excel').send_keys('C:/Users/13209/Desktop/测试脚本/模板文件/tepl.xlsx')
    time.sleep(3)
    driver.find_element_by_xpath('//div[@class="layui-layer-btn layui-layer-btn-c"]/a[1]').click()
    num+=1000
driver.quit()